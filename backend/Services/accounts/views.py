from django.contrib.auth import authenticate, login
from django.shortcuts import render, redirect
from django.contrib.auth.forms import AuthenticationForm
from django.contrib import messages
from .forms import CustomUserForm
from core.models import Hostel_Management  # adjust to actual model name
from django.db import models


def user_login(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                # Redirect to a dashboard based on role
                if user.role == 'student':
                    return redirect('student_dashboard')
                elif user.role == 'warden':
                    return redirect('warden_dashboard')
                elif user.role == 'chief_warden':
                    return redirect('chief_warden_dashboard')
                elif user.role == 'caretaker':
                    return redirect('caretaker_dashboard')
                elif user.role == 'guard':
                    return redirect('guard_dashboard')
            else:
                messages.error(request, 'Invalid credentials')
        else:
            messages.error(request, 'Invalid form submission')
    else:
        form = AuthenticationForm()
    return render(request, 'accounts/login.html', {'form': form})


from django.shortcuts import render
from django.contrib.auth.decorators import login_required

@login_required
def warden_dashboard(request):
    return render(request, 'accounts/warden_dashboard.html')

@login_required
def student_dashboard(request):
    return render(request, 'accounts/student_dashboard.html')

@login_required
def chief_warden_dashboard(request):
    return render(request, 'accounts/chief_warden_dashboard.html')

@login_required
def caretaker_dashboard(request):
    return render(request, 'accounts/caretaker_dashboard.html')

@login_required
def guard_dashboard(request):
    return render(request, 'accounts/guard_dashboard.html')


from django.contrib.auth import logout
from django.shortcuts import redirect

def user_logout(request):
    logout(request)
    return redirect('login')




# --------------------------user creation hostel HostelManagement---------------------------------


def register_user(request):
    if request.method == 'POST':
        form = CustomUserForm(request.POST)
        if form.is_valid():
            user = form.save()
            # Create related model data based on role
            role = user.role

            if role in ['warden', 'chief_warden', 'caretaker', 'guard']:
                Hostel_Management.objects.create(user=user)  # Add any other required fields

            return redirect('user_success')  # Or a dashboard or confirmation
    else:
        form = CustomUserForm()

    return render(request, 'accounts/register_user.html', {'form': form})


from django.http import HttpResponse

def user_success(request):
    return HttpResponse("User created successfully.")



# ___________________________________________________view for creating student__________________________

from django.shortcuts import render, redirect
from django.contrib import messages
from accounts.forms import StudentUserCreationForm

def register_student(request):
    if request.method == 'POST':
        form = StudentUserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Student account created successfully.")
            return redirect('student-register')  # Or another success page
        else:
            messages.error(request, "There was an error in the form. Please correct it.")
    else:
        form = StudentUserCreationForm()
    
    return render(request, 'accounts/registerstudent.html', {'form': form})



# ___________________________________________________view for creating hostel management__________________________



from django.shortcuts import render, redirect
from django.contrib import messages
from accounts.forms import HostelManagerUserCreationForm

def register_hostel_manager(request):
    if request.method == 'POST':
        form = HostelManagerUserCreationForm(request.POST)
        if form.is_valid():
            # Save the form to create the CustomUser object
            user = form.save(commit=False)

            # Set the user as staff
            user.is_staff = True

            # Save the user object with staff status
            user.save()

            messages.success(request, "Hostel manager account created successfully.")
            return redirect('hostel-manager-register')  # or another page
        else:
            messages.error(request, "There was an error in the form. Please correct it.")
    else:
        form = HostelManagerUserCreationForm()

    return render(request, 'accounts/register_hostel_manager.html', {'form': form})




# -----------bulk upload -----------------


from django.shortcuts import render
from django.http import HttpResponse
from django.contrib import messages
from django.db import transaction
import openpyxl
import tempfile

from core.models import Student
from .models import CustomUser  # Adjust the import path if needed

def bulk_upload_students(request):
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        wb = openpyxl.load_workbook(file)
        sheet = wb.active

        headers = [cell.value for cell in sheet[1]]
        failed_rows = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            try:
                with transaction.atomic():
                    # Normalize gender input (default to 'm' if invalid/missing)
                    gender_input = str(data.get('gender', 'm')).strip().lower()

                    gender = gender_input if gender_input in ['m', 'f', 'o'] else 'm'

                    user = CustomUser.objects.create_user(
                        username=data['username'],
                        password=data['password'],
                        email=data['email'],
                        role='student',
                    )

                    Student.objects.create(
                        user=user,
                        name=data['name'],
                        enroll_number=int(data['enroll_number']),
                        email=data['email'],
                        student_contact=int(data['student_contact']),
                        parent_contact1=int(data['parent_contact1']),
                        parent_contact2=int(data['parent_contact2']),
                        cg=float(data['cg']),
                        admn_year=int(data['admn_year']),
                        gender=gender,
                    )
            except Exception as e:
                if 'user' in locals() and user.pk:
                    user.delete()
                failed_data = list(row) + [str(e)]
                failed_rows.append(failed_data)

        if failed_rows:
            output_wb = openpyxl.Workbook()
            output_ws = output_wb.active
            output_ws.title = "Failed Records"
            output_ws.append(headers + ['error'])

            for row in failed_rows:
                output_ws.append([cell for cell in row])  # Avoid Cell objects

            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                output_wb.save(tmp.name)
                tmp.seek(0)
                response = HttpResponse(
                    tmp.read(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = 'attachment; filename=failed_uploads.xlsx'
                return response

        messages.success(request, "Bulk student upload completed successfully.")
        return render(request, 'accounts/bulk_upload.html')

    return render(request, 'accounts/bulk_upload.html')



# -------------------------edit student----------------------------




from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from core.models import Student
from accounts.forms import StudentEditForm  # You’ll need to create this form
from django.db.models import Q
from django.core.paginator import Paginator

@login_required
def edit_student(request, student_id):
    student = get_object_or_404(Student, id=student_id)

    if request.method == 'POST':
        form = StudentEditForm(request.POST, instance=student)
        if form.is_valid():
            form.save()
            messages.success(request, 'Student details updated successfully.')
            return redirect('search-students')
    else:
        form = StudentEditForm(instance=student)

    return render(request, 'accounts/edit_student.html', {'form': form, 'student': student})




# -------------------------search bar----------------------------



from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import render, redirect
from django.db.models import Q
from django.core.paginator import Paginator
from core.models import Student  # Adjust as needed

# @login_required
# def search_students(request):
#     if not request.user.is_staff:
#         messages.error(request, f"Access denied for user '{request.user.username}'. Staff privileges are required.")
#         return redirect('login')

#     query = request.GET.get('q', '')
#     student_list = Student.objects.all()

#     if query:
#         student_list = student_list.filter(
#             Q(name__icontains=query) |
#             Q(enroll_number__icontains=query) |
#             Q(email__icontains=query)
#         )

#     paginator = Paginator(student_list.order_by('name'), 10)
#     page_number = request.GET.get('page')
#     students = paginator.get_page(page_number)

#     return render(request, 'accounts/search_students.html', {
#         'students': students,
#         'query': query,
#         'active_page': 'Manage students',  # This matches the sidebar label
#     })


from django.utils.timezone import now
from datetime import datetime

@login_required
def search_students(request):
    if not request.user.is_staff:
        messages.error(request, f"Access denied for user '{request.user.username}'. Staff privileges are required.")
        return redirect('login')

    query = request.GET.get('q', '')
    gender_filter = request.GET.get('gender', '')
    hostel_filter = request.GET.get('hostel', '')
    year_filter = request.GET.get('admn_year', '')

    student_list = Student.objects.all()

    if query:
        student_list = student_list.filter(
            Q(name__icontains=query) |
            Q(enroll_number__icontains=query) |
            Q(email__icontains=query)
        )
    if gender_filter:
        student_list = student_list.filter(gender=gender_filter)
    if hostel_filter:
        student_list = student_list.filter(room__hostel__name__icontains=hostel_filter)
    if year_filter:
        student_list = student_list.filter(admn_year=year_filter)

    paginator = Paginator(student_list.order_by('name'), 10)
    page_number = request.GET.get('page')
    students = paginator.get_page(page_number)

    current_year = now().year
    year_range = [str(y) for y in range(current_year - 4, current_year + 1)]

    return render(request, 'accounts/search_students.html', {
        'students': students,
        'query': query,
        'gender_filter': gender_filter,
        'hostel_filter': hostel_filter,
        'year_filter': year_filter,
        'year_range': year_range,
        'active_page': 'Manage students',
    })



# --------------------------- room view -----------------------------



from django.shortcuts import render, redirect
from django.core.paginator import Paginator
from django.db.models import Q
from core.models import HostelRoom, Hostels
from django.contrib.auth.decorators import login_required

@login_required
def hostel_room_list_view(request):
    # # Check if user is either Warden or Chief Warden
    # if not request.user.has_perm('core.view_hostelroom'):
    #     return redirect('login')  # or any other view for unauthorized access

    if request.method == 'POST':
        room_ids = request.POST.getlist('room_id')
        for room_id in room_ids:
            room = HostelRoom.objects.get(id=room_id)
            new_value = request.POST.get(f'show_{room_id}') == 'on'
            if room.show != new_value:
                room.show = new_value
                room.save()

        return redirect('room_management')  # After saving, redirect to the same view

    # Filters from GET request
    query = request.GET.get('q', '')
    selected_hostel = request.GET.get('hostel', '')
    occupancy_filter = request.GET.get('occupancy', '')
    capacity_filter = request.GET.get('capacity', '')  # New capacity filter

    room_list = HostelRoom.objects.all()

    # Apply filters
    if query:
        room_list = room_list.filter(
            Q(room_number__icontains=query) |
            Q(hostel__name__icontains=query)
        )
    if selected_hostel:
        room_list = room_list.filter(hostel__id=selected_hostel)
    if occupancy_filter:
        if occupancy_filter == 'full':
            room_list = room_list.filter(occupancy=models.F('capacity'))
        elif occupancy_filter == 'available':
            room_list = room_list.filter(occupancy=0)
        elif occupancy_filter == 'partial':
            room_list = room_list.filter(occupancy__gt=0, occupancy__lt=models.F('capacity'))
    if capacity_filter:
        room_list = room_list.filter(capacity=capacity_filter)  # Filter by capacity

    paginator = Paginator(room_list.order_by('room_number'), 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    hostels = Hostels.objects.all()

    return render(request, 'accounts/room_management.html', {
        'page_obj': page_obj,
        'query': query,
        
        'hostels': hostels,
        'selected_hostel': selected_hostel,
        'occupancy_filter': occupancy_filter,
        'capacity_filter': capacity_filter,  # Pass capacity filter to the template
        'active_page': 'Manage rooms' 
    })


# ---------------------------edit room -------------------------

from django.shortcuts import get_object_or_404
from django import forms

class HostelRoomForm(forms.ModelForm):
    class Meta:
        model = HostelRoom
        fields = ['room_number', 'hostel', 'capacity', 'occupancy', 'show']

@login_required
def edit_room_view(request, pk):
    # if not request.user.has_perm('core.change_hostelroom'):
    #     return redirect('home')

    room = get_object_or_404(HostelRoom, pk=pk)

    if request.method == 'POST':
        form = HostelRoomForm(request.POST, instance=room)
        if form.is_valid():
            form.save()
            return redirect('room_management')  # Redirect back to room list
    else:
        form = HostelRoomForm(instance=room)

    return render(request, 'accounts/edit_room.html', {'form': form, 'room': room})













# # --------------------upload room data-----------------------------



from core.models import HostelRoom, Hostels, InventoryForm  # Make sure InventoryForm is imported
import openpyxl
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import render
from django.contrib import messages
import tempfile


def bulk_upload_rooms(request):
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        wb = openpyxl.load_workbook(file)
        sheet = wb.active

        headers = [cell.value for cell in sheet[1]]
        failed_rows = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            try:
                with transaction.atomic():
                    hostel = Hostels.objects.get(name=data['hostel'])

                    room = HostelRoom.objects.create(
                        room_number=int(data['room_number']),
                        hostel=hostel,
                        capacity=int(data.get('capacity', 2)),
                        floor=int(data.get('floor', 0)),
                        level=int(data.get('level', 0)),
                        balcony=bool(int(data.get('balcony', 0))),
                        sunny=bool(int(data.get('sunny', 0))),
                        show=bool(int(data.get('show', 1))),
                    )

                    # Create empty inventory form for this room
                    InventoryForm.objects.create(hostel_room=room)

            except Exception as e:
                failed_data = list(row) + [str(e)]
                failed_rows.append(failed_data)

        if failed_rows:
            output_wb = openpyxl.Workbook()
            output_ws = output_wb.active
            output_ws.title = "Failed Records"
            output_ws.append(headers + ['error'])

            for row in failed_rows:
                output_ws.append(row)

            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                output_wb.save(tmp.name)
                tmp.seek(0)
                response = HttpResponse(
                    tmp.read(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = 'attachment; filename=failed_room_uploads.xlsx'
                return response

        messages.success(request, "Bulk room upload completed successfully.")
        return render(request, 'accounts/bulk_upload_rooms.html')

    return render(request, 'accounts/bulk_upload_rooms.html')


# ----------------------------------public room list student---------------------------------------------

from django.shortcuts import render
from core.models import HostelRoom  # Replace with your actual model name if different

def available_rooms_view(request):
    rooms = HostelRoom.objects.filter(show=True).filter(capacity__gt=models.F('occupancy'))
    return render(request, 'accounts/available_rooms.html', {'rooms': rooms})


